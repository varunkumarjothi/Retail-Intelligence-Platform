"""
Automated Report Generator — Retail Intelligence Platform
==========================================================
Generates a multi-sheet Excel business report + Matplotlib charts.
Author: Varun Kumar Jothi
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import os, sys, sqlite3
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config.settings import DB_PATH, REPORTS_DIR, REPORT_TITLE
from sql.analytics import (overall_kpis, revenue_trend_monthly, top_products,
                            region_segment_matrix, sales_rep_leaderboard,
                            discount_impact_analysis)

os.makedirs(REPORTS_DIR, exist_ok=True)
COLORS = ['#1A5276','#2E86C1','#5DADE2','#85C1E9','#154360','#1F618D']


def save_chart(fig, name):
    path = os.path.join(REPORTS_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Chart: {path}")
    return path


def chart_revenue_trend(df):
    df['period'] = df['order_year'].astype(str) + '-' + df['order_month'].astype(str).str.zfill(2)
    fig, ax1 = plt.subplots(figsize=(14, 5))
    ax2 = ax1.twinx()
    ax1.fill_between(df['period'], df['revenue'], alpha=0.25, color=COLORS[0])
    ax1.plot(df['period'], df['revenue'], color=COLORS[0], linewidth=2.5, marker='o', markersize=4, label='Revenue')
    ax2.plot(df['period'], df['avg_margin_pct'], color='#E74C3C', linewidth=2, linestyle='--', marker='s', markersize=4, label='Margin %')
    ax1.set_title('Monthly Revenue & Profit Margin Trend', fontsize=14, fontweight='bold', pad=12)
    ax1.set_xlabel('Month')
    ax1.set_ylabel('Revenue (Rs.)', color=COLORS[0])
    ax2.set_ylabel('Profit Margin %', color='#E74C3C')
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'Rs.{x/1e5:.1f}L'))
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1+lines2, labels1+labels2, loc='upper left')
    ax1.grid(axis='y', linestyle='--', alpha=0.4)
    ax1.xaxis.set_major_locator(plt.MaxNLocator(12))
    plt.xticks(rotation=45, ha='right', fontsize=8)
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    return save_chart(fig, 'chart_01_revenue_trend.png')


def chart_category_breakdown(df_top):
    cat = df_top.groupby('product_category')[['total_revenue','total_profit']].sum()
    margins = df_top.groupby('product_category')['avg_margin_pct'].mean().sort_values(ascending=False)

    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    # Pie chart with explode for small slices
    explode = [0.05] * len(cat)
    wedges, texts, autotexts = axes[0].pie(
        cat['total_revenue'],
        labels=None,
        colors=COLORS[:len(cat)],
        startangle=140,
        explode=explode,
        wedgeprops={'edgecolor': 'white', 'linewidth': 2},
        autopct=lambda pct: f'{pct:.1f}%' if pct > 5 else '',
        pctdistance=0.75
    )
    for autotext in autotexts:
        autotext.set_fontsize(11)
        autotext.set_fontweight('bold')

    # Show all values in legend including small ones
    legend_labels = [f"{cat.index[i]} ({cat['total_revenue'].iloc[i]/cat['total_revenue'].sum()*100:.1f}%)"
                     for i in range(len(cat))]
    axes[0].legend(legend_labels, loc='lower center',
                   bbox_to_anchor=(0.5, -0.12),
                   ncol=2, fontsize=10)
    axes[0].set_title('Revenue Share by Category', fontweight='bold', fontsize=13, pad=15)
    # Horizontal bar chart
    bars = axes[1].barh(margins.index, margins.values,
                        color=COLORS[:len(margins)], height=0.5)
    for bar in bars:
        axes[1].text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                     f'{bar.get_width():.1f}%', va='center',
                     fontsize=11, fontweight='bold')
    axes[1].set_title('Average Margin by Category', fontweight='bold', fontsize=13)
    axes[1].set_xlabel('Profit Margin %')
    axes[1].set_xlim(0, margins.max() + 12)
    axes[1].grid(axis='x', linestyle='--', alpha=0.4)
    axes[1].spines['top'].set_visible(False)
    axes[1].spines['right'].set_visible(False)

    plt.tight_layout()
    return save_chart(fig, 'chart_02_category.png')

def chart_sales_rep(df_rep):
    x = np.arange(len(df_rep))
    fig, ax = plt.subplots(figsize=(11, 6))
    b1 = ax.bar(x - 0.2, df_rep['total_revenue']/1e5, 0.35, label='Revenue (L)', color=COLORS[0])
    b2 = ax.bar(x + 0.2, df_rep['total_profit']/1e5,  0.35, label='Profit (L)',  color=COLORS[2])
    ax.set_xticks(x)
    ax.set_xticklabels(df_rep['sales_rep'], fontsize=11)
    ax.set_title('Sales Rep Performance — Revenue vs Profit', fontsize=14, fontweight='bold', pad=12)
    ax.set_ylabel('Amount (Lakhs)')
    ax.legend()
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    for b in list(b1)+list(b2):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.1,
                f'{b.get_height():.1f}L', ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    return save_chart(fig, 'chart_03_sales_rep.png')


def chart_region_heatmap(df_region):
    pivot = df_region.pivot(index='region', columns='segment', values='revenue').fillna(0)
    fig, ax = plt.subplots(figsize=(10, 5))
    im = ax.imshow(pivot.values, cmap='Blues', aspect='auto')
    ax.set_xticks(range(len(pivot.columns)))
    ax.set_xticklabels(pivot.columns, fontsize=11)
    ax.set_yticks(range(len(pivot.index)))
    ax.set_yticklabels(pivot.index, fontsize=11)
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            val = pivot.values[i, j]
            ax.text(j, i, f'Rs.{val/1e5:.1f}L', ha='center', va='center',
                    fontsize=10, color='white' if val > pivot.values.max()*0.6 else 'black')
    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label('Revenue (Rs.)', fontsize=10)
    cbar.formatter = mticker.FuncFormatter(lambda x, _: f'Rs.{x/1e5:.1f}L')
    cbar.update_ticks()
    ax.set_title('Revenue Heatmap — Region × Segment', fontsize=14, fontweight='bold', pad=12)
    plt.tight_layout()
    return save_chart(fig, 'chart_04_region_heatmap.png')

def chart_forecast(forecast_csv):
    if not os.path.exists(forecast_csv):
        return None
    df_f = pd.read_csv(forecast_csv)
    fig, ax = plt.subplots(figsize=(12, 6))

    bars = ax.bar(range(len(df_f)), df_f['forecasted_revenue']/1e5,
                  color=COLORS[1], alpha=0.85, width=0.6)
    ax.plot(range(len(df_f)), df_f['forecasted_revenue']/1e5,
            color=COLORS[0], marker='o', linewidth=2.5, markersize=8, zorder=5)

    # Add value labels on top of each bar
    for i, bar in enumerate(bars):
        ax.text(bar.get_x() + bar.get_width()/2,
                bar.get_height() + 0.02,
                f'Rs.{df_f["forecasted_revenue"].iloc[i]/1e5:.2f}L',
                ha='center', va='bottom', fontsize=8, fontweight='bold')

    ax.set_xticks(range(len(df_f)))
    ax.set_xticklabels([str(w)[:10] for w in df_f['week']],
                       rotation=30, ha='right', fontsize=9)

    # Start y-axis close to min value so differences are visible
    min_val = df_f['forecasted_revenue'].min()/1e5
    max_val = df_f['forecasted_revenue'].max()/1e5
    ax.set_ylim(min_val * 0.98, max_val * 1.05)

    ax.set_title('8-Week Revenue Forecast (ML Model)', fontsize=14,
                 fontweight='bold', pad=12)
    ax.set_ylabel('Forecasted Revenue (Lakhs)')
    ax.set_xlabel('Week')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    plt.tight_layout()
    return save_chart(fig, 'chart_05_forecast.png')

def generate_excel_report(kpis, df_trend, df_top, df_region, df_rep, df_discount):
    path = os.path.join(REPORTS_DIR, 'retail_intelligence_report.xlsx')
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        pd.DataFrame([kpis]).to_excel(writer, sheet_name='Overall KPIs', index=False)
        df_trend.to_excel(writer,    sheet_name='Monthly Trend',    index=False)
        df_top.to_excel(writer,      sheet_name='Top Products',      index=False)
        df_region.to_excel(writer,   sheet_name='Region Analysis',   index=False)
        df_rep.to_excel(writer,      sheet_name='Sales Rep KPIs',    index=False)
        df_discount.to_excel(writer, sheet_name='Discount Analysis', index=False)
        fc = os.path.join(os.path.dirname(DB_PATH), 'forecast.csv')
        if os.path.exists(fc):
            pd.read_csv(fc).to_excel(writer, sheet_name='8-Week Forecast', index=False)

        # Auto-fit all columns in all sheets
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Style header row
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1A5276',
                                       end_color='1A5276',
                                       fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 4

    print(f"  Excel report: {path}")
    return path

def run_report():
    print("=" * 55)
    print("  REPORT GENERATOR")
    print("=" * 55)

    kpis        = overall_kpis()
    df_trend    = revenue_trend_monthly()
    df_top      = top_products(20)
    df_region   = region_segment_matrix()
    df_rep      = sales_rep_leaderboard()
    df_discount = discount_impact_analysis()
    fc_csv      = os.path.join(os.path.dirname(DB_PATH), 'forecast.csv')

    print("\nGenerating charts...")
    chart_revenue_trend(df_trend.copy())
    chart_category_breakdown(df_top.copy())
    chart_sales_rep(df_rep.copy())
    chart_region_heatmap(df_region.copy())
    chart_forecast(fc_csv)

    print("\nGenerating Excel report...")
    generate_excel_report(kpis, df_trend, df_top, df_region, df_rep, df_discount)

    print(f"\n  Key KPIs:")
    print(f"    Total Revenue : Rs.{kpis['total_revenue']:,.0f}")
    print(f"    Total Profit  : Rs.{kpis['total_profit']:,.0f}")
    print(f"    Avg Margin    : {kpis['avg_margin_pct']}%")
    print(f"    Total Orders  : {kpis['total_orders']:,}")
    print("\n  Report generation complete!")


if __name__ == "__main__":
    run_report()